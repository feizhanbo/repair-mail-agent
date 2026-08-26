from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import logging
import re
import time
import zipfile
from html import unescape
from pathlib import PurePath
from typing import Any
from uuid import uuid4
from xml.etree import ElementTree

from bs4 import BeautifulSoup
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.ai.prompts import ATTACHMENT_TEXT, ATTACHMENT_VISUAL
from app.integrations.ai_provider import AiProviderError
from app.integrations.llm_gateway import LlmTask, invoke_structured, llm_task_configured
from app.models import EmailAttachment
from app.services.attachment_precheck import (
    ARCHIVE_CONTENT_TYPES,
    detect_archive_format,
    engineering_reference_metadata,
)
from app.services.common import utcnow
from app.services.logging_safety import safe_error_code
from app.services.storage import (
    StorageConfigurationError, StorageUploadError, download_oss_object_bytes,
    generate_presigned_url_for_object,
)


SUPPORTED_ATTACHMENT_TYPES = {"docx", "xlsx", "csv", "txt", "prc", "html", "image", "pdf"}
_file_parse_semaphore = asyncio.Semaphore(max(1, settings.FILE_PARSE_CONCURRENCY))
logger = logging.getLogger(__name__)


class AttachmentParseJson(BaseModel):
    file_type: str
    summary: str = ""
    key_points: list[str] = Field(default_factory=list)
    extracted_fields: dict[str, Any] = Field(default_factory=dict)
    extracted_items: list[dict[str, Any]] = Field(default_factory=list)
    raw_text: str = ""
    warnings: list[str] = Field(default_factory=list)
    truncated: bool = False

    @field_validator("summary", "raw_text", mode="before")
    @classmethod
    def normalize_text(cls, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        return json.dumps(value, ensure_ascii=False, default=str)

    @field_validator("key_points", "warnings", mode="before")
    @classmethod
    def normalize_string_list(cls, value: Any) -> list[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [item if isinstance(item, str) else json.dumps(item, ensure_ascii=False, default=str) for item in value]
        return [value if isinstance(value, str) else json.dumps(value, ensure_ascii=False, default=str)]

    @field_validator("extracted_fields", mode="before")
    @classmethod
    def normalize_fields(cls, value: Any) -> dict[str, Any]:
        return value if isinstance(value, dict) else {}

    @field_validator("extracted_items", mode="before")
    @classmethod
    def normalize_items(cls, value: Any) -> list[dict[str, Any]]:
        if isinstance(value, dict):
            return [value]
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
        return []

    @field_validator("truncated", mode="before")
    @classmethod
    def normalize_truncated(cls, value: Any) -> bool:
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "是"}
        return bool(value)


def attachment_type(attachment: EmailAttachment) -> str | None:
    content_type = (attachment.content_type or "").lower().split(";", 1)[0].strip()
    suffix = PurePath(attachment.file_name or "").suffix.lower().lstrip(".")
    if content_type.startswith("image/"):
        return "image"
    if content_type == "application/pdf" or suffix == "pdf":
        return "pdf"
    if suffix in {"docx", "xlsx", "csv", "txt", "prc", "html", "htm"}:
        return "html" if suffix == "htm" else suffix
    if content_type in {"text/plain", "application/json", "application/xml", "text/xml"}:
        return "txt"
    if content_type in {"text/csv", "application/csv"}:
        return "csv"
    if content_type in {"text/html", "application/xhtml+xml"}:
        return "html"
    return None


def _qwen_configured(*, visual: bool = False) -> bool:
    return llm_task_configured(
        LlmTask.ATTACHMENT_VISUAL_PARSE if visual else LlmTask.ATTACHMENT_TEXT_PARSE
    )


async def _file_io(func, *args, **kwargs):
    async with _file_parse_semaphore:
        return await asyncio.to_thread(func, *args, **kwargs)


def _qwen_retryable(error: AiProviderError) -> bool:
    code = str(error).upper()
    return any(
        marker in code
        for marker in ("TIMEOUT", "HTTP_429", "HTTP_5", "OUTPUT_NOT_JSON", "OUTPUT_SCHEMA_INVALID")
    )


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _compact_lines(text: str, *, limit: int = 12) -> list[str]:
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    return [line for line in lines if line][:limit]


def _local_summary(text: str) -> tuple[str, list[str]]:
    lines = _compact_lines(text, limit=8)
    return ("\n".join(lines)[:1000], lines[:5])


def _truncate_text(text: str) -> tuple[str, bool]:
    limit = max(1000, settings.ATTACHMENT_TEXT_MAX_CHARS)
    if len(text) <= limit:
        return text, False
    return text[:limit], True


def _extract_txt(content: bytes) -> str:
    return _decode_text(content)


def _extract_prc(content: bytes) -> str:
    if not content or content.count(b"\x00") / len(content) > 0.02:
        raise ValueError("PRC_TEXT_UNRECOGNIZED")
    text = _decode_text(content)
    sample = text[:10000]
    printable = sum(character.isprintable() or character in "\r\n\t" for character in sample)
    if sample and printable / len(sample) < 0.9:
        raise ValueError("PRC_TEXT_UNRECOGNIZED")
    return text


def _extract_html(content: bytes) -> str:
    html = _decode_text(content)
    return unescape(BeautifulSoup(html, "lxml").get_text("\n"))


def _extract_csv(content: bytes) -> str:
    text = _decode_text(content)
    reader = csv.reader(io.StringIO(text))
    rows: list[str] = []
    for index, row in enumerate(reader):
        if index >= 200:
            rows.append("... truncated after 200 rows ...")
            break
        rows.append(" | ".join(str(cell).strip() for cell in row))
    return "\n".join(rows)


def _validate_zip_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
    except Exception as exc:
        raise ValueError("ARCHIVE_INVALID") from exc
    if len(members) > 1000:
        raise ValueError("ARCHIVE_MEMBER_LIMIT_EXCEEDED")
    expanded_size = sum(member.file_size for member in members)
    compressed_size = sum(max(1, member.compress_size) for member in members)
    if expanded_size > 100 * 1024 * 1024:
        raise ValueError("ARCHIVE_EXPANDED_SIZE_EXCEEDED")
    if expanded_size / max(1, compressed_size) > 100:
        raise ValueError("ARCHIVE_COMPRESSION_RATIO_EXCEEDED")


def _extract_docx(content: bytes) -> str:
    _validate_zip_archive(content)
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            xml = archive.read("word/document.xml")
    except Exception as exc:
        raise ValueError("DOCX_TEXT_EXTRACT_FAILED") from exc
    root = ElementTree.fromstring(xml)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs: list[str] = []
    for paragraph in root.iter(f"{namespace}p"):
        parts = [node.text or "" for node in paragraph.iter(f"{namespace}t")]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _extract_xlsx(
    content: bytes, *, max_sheets: int = 10, max_rows: int = 200, max_columns: int = 40
) -> str:
    _validate_zip_archive(content)
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError("XLSX_READER_NOT_AVAILABLE") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("XLSX_TEXT_EXTRACT_FAILED") from exc
    blocks: list[str] = []
    for sheet in workbook.worksheets[:max_sheets]:
        blocks.append(f"# sheet: {sheet.title}")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index > max_rows:
                blocks.append(f"... truncated after {max_rows} rows ...")
                break
            values = ["" if value is None else str(value) for value in row[:max_columns]]
            if any(value.strip() for value in values):
                blocks.append(" | ".join(values))
    return "\n".join(blocks)


def _pdf_page_count(content: bytes) -> int | None:
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(content))
        return len(reader.pages)
    except Exception:
        matches = re.findall(rb"/Type\s*/Page\b", content)
        return len(matches) or None


def _extract_pdf_text(content: bytes, *, max_pages: int) -> tuple[str, int | None]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return "", _pdf_page_count(content)
    try:
        reader = PdfReader(io.BytesIO(content))
        if reader.is_encrypted:
            raise ValueError("PDF_ENCRYPTED")
        page_count = len(reader.pages)
        texts = [(reader.pages[index].extract_text() or "") for index in range(min(page_count, max_pages))]
        return "\n".join(part.strip() for part in texts if part.strip()), page_count
    except ValueError:
        raise
    except Exception:
        raise ValueError("PDF_READ_FAILED")


def render_pdf_pages(content: bytes, *, max_pages: int) -> tuple[list[bytes], int]:
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise ValueError("PDF_RENDERER_NOT_AVAILABLE") from exc
    try:
        document = fitz.open(stream=content, filetype="pdf")
        if document.needs_pass:
            raise ValueError("PDF_ENCRYPTED")
        page_count = document.page_count
        pages = [
            document.load_page(index).get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False).tobytes("png")
            for index in range(min(page_count, max_pages))
        ]
        document.close()
        return pages, page_count
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("PDF_RENDER_FAILED") from exc


def _image_dimensions(content: bytes) -> tuple[int, int] | None:
    if content.startswith(b"\x89PNG\r\n\x1a\n") and len(content) >= 24:
        return int.from_bytes(content[16:20], "big"), int.from_bytes(content[20:24], "big")
    if content[:6] in {b"GIF87a", b"GIF89a"} and len(content) >= 10:
        return int.from_bytes(content[6:8], "little"), int.from_bytes(content[8:10], "little")
    if content.startswith(b"\xff\xd8"):
        index = 2
        start_of_frame = set(range(0xC0, 0xC4)) | set(range(0xC5, 0xC8)) | set(range(0xC9, 0xCC)) | set(range(0xCD, 0xD0))
        while index + 9 < len(content):
            if content[index] != 0xFF:
                index += 1
                continue
            marker = content[index + 1]
            index += 2
            if marker in {0xD8, 0xD9}:
                continue
            if index + 2 > len(content):
                break
            segment_size = int.from_bytes(content[index:index + 2], "big")
            if marker in start_of_frame and index + 7 < len(content):
                height = int.from_bytes(content[index + 3:index + 5], "big")
                width = int.from_bytes(content[index + 5:index + 7], "big")
                return width, height
            index += max(segment_size, 2)
    return None


def _fallback_json(file_type: str, text: str, *, warnings: list[str], truncated: bool) -> AttachmentParseJson:
    summary, key_points = _local_summary(text)
    return AttachmentParseJson(
        file_type=file_type,
        summary=summary,
        key_points=key_points,
        raw_text=text,
        warnings=warnings,
        truncated=truncated,
    )


async def _invoke_qwen(
    session: AsyncSession,
    *,
    attachment: EmailAttachment,
    call_type: str,
    visual: bool,
    input_payload: dict[str, Any],
    invoke,
) -> AttachmentParseJson:
    if not _qwen_configured(visual=visual):
        raise AiProviderError("QWEN_VL_NOT_CONFIGURED" if visual else "QWEN_TEXT_NOT_CONFIGURED")

    max_attempts = 1
    last_error: AiProviderError | None = None
    for attempt in range(1, max_attempts + 1):
        logger.info(
            "Attachment AI call started",
            extra={
                "event": "ai_call_started", "call_type": call_type,
                "attachment_id": attachment.id, "email_id": attachment.email_id,
                "attempt": attempt,
            },
        )
        try:
            completion = await invoke()
            if hasattr(session, "add"):
                from app.services.ai import persist_ai_log

                await persist_ai_log(
                    session,
                    trace_id=getattr(completion, "trace_id", uuid4().hex),
                    call_type=call_type,
                    input_payload=input_payload,
                    request_payload=getattr(completion, "request_payload", None),
                    output_payload=getattr(completion, "response_payload", None),
                    parsed=completion.parsed,
                    latency_ms=getattr(completion, "latency_ms", None),
                    input_summary=f"attachment_id={attachment.id}; file_type={input_payload.get('file_type')}",
                    output_summary=f"attachment_type={completion.parsed.file_type}; warnings={len(completion.parsed.warnings)}",
                    email_id=attachment.email_id,
                    attachment_id=attachment.id,
                    provider_name=getattr(completion, "provider_name", None) or "unknown",
                    model_name=getattr(completion, "model_name", None) or "unknown",
                    prompt_version=(ATTACHMENT_VISUAL if visual else ATTACHMENT_TEXT).version,
                    prompt_hash=(ATTACHMENT_VISUAL if visual else ATTACHMENT_TEXT).content_hash,
                    route_name=getattr(completion, "route_name", None),
                    route_attempt=int(getattr(completion, "route_attempt", 1)),
                    fallback_used=bool(getattr(completion, "fallback_used", False)),
                    attempt_count=int(getattr(completion, "route_attempt", attempt)),
                )
            return completion.parsed
        except AiProviderError as exc:
            last_error = exc
            error_code = safe_error_code(exc, "QWEN_CALL_FAILED")
            if hasattr(session, "add"):
                from app.services.ai import persist_ai_log

                await persist_ai_log(
                    session,
                    trace_id=getattr(exc, "trace_id", uuid4().hex),
                    call_type=call_type,
                    input_payload=input_payload,
                    request_payload=getattr(exc, "request_payload", None),
                    output_payload=(
                        getattr(exc, "response_payload", None)
                        or {"raw_output": getattr(exc, "raw_output", None)}
                    ),
                    parsed=None,
                    latency_ms=getattr(exc, "latency_ms", None),
                    input_summary=f"attachment_id={attachment.id}; file_type={input_payload.get('file_type')}",
                    output_summary=error_code,
                    email_id=attachment.email_id,
                    attachment_id=attachment.id,
                    provider_name=str(getattr(exc, "route_name", "unknown")),
                    model_name=str(getattr(exc, "model_name", "unknown")),
                    prompt_version=(ATTACHMENT_VISUAL if visual else ATTACHMENT_TEXT).version,
                    prompt_hash=(ATTACHMENT_VISUAL if visual else ATTACHMENT_TEXT).content_hash,
                    route_name=getattr(exc, "route_name", None),
                    route_attempt=int(getattr(exc, "route_attempt", attempt)),
                    fallback_used=int(getattr(exc, "route_attempt", attempt)) > 1,
                    attempt_count=int(getattr(exc, "route_attempt", attempt)),
                    error_message=error_code,
                )
            if attempt >= max_attempts or not _qwen_retryable(exc):
                break
            await asyncio.sleep(settings.AI_RETRY_BASE_DELAY_SECONDS * (2 ** (attempt - 1)))
    assert last_error is not None
    raise last_error


async def _qwen_text_parse(
    session: AsyncSession,
    *,
    attachment: EmailAttachment,
    file_type: str,
    file_name: str,
    text: str,
    warnings: list[str],
    truncated: bool,
) -> AttachmentParseJson:
    if not _qwen_configured(visual=False):
        raise AiProviderError("QWEN_API_KEY_NOT_CONFIGURED")
    summary, key_points = _local_summary(text)
    prompt = (
        "请将附件内容解析为维修邮件附件级 JSON。只能输出 JSON，字段固定为 "
        "file_type, summary, key_points, extracted_fields, extracted_items, raw_text, warnings, truncated。\n"
        f"file_name={file_name}\nfile_type={file_type}\ntruncated={truncated}\n"
        f"local_summary={summary}\nlocal_key_points={key_points}\ncontent:\n{text}"
    )
    messages = [
            {
                "role": "system",
                "content": ATTACHMENT_TEXT.system,
            },
            {"role": "user", "content": prompt},
        ]
    parsed = await _invoke_qwen(
        session,
        attachment=attachment,
        call_type="attachment_text_parse",
        visual=False,
        input_payload={"file_type": file_type, "text": text, "truncated": truncated},
        invoke=lambda: invoke_structured(
            task=LlmTask.ATTACHMENT_TEXT_PARSE,
            messages=messages,
            response_model=AttachmentParseJson,
            temperature=0.1,
        ),
    )
    parsed.file_type = parsed.file_type or file_type
    parsed.warnings = [*warnings, *(parsed.warnings or [])]
    parsed.truncated = bool(parsed.truncated or truncated)
    return parsed


async def _qwen_visual_parse(
    session: AsyncSession,
    *,
    attachment: EmailAttachment,
    file_type: str,
    file_name: str,
    urls: list[str],
    warnings: list[str],
    truncated: bool,
) -> AttachmentParseJson:
    if not _qwen_configured(visual=True):
        raise AiProviderError("QWEN_API_KEY_NOT_CONFIGURED")
    prompt = (
        f"{ATTACHMENT_VISUAL.system}\n请识别并提取图片或 PDF 附件中的维修报修相关信息。只能输出 JSON，字段固定为 "
        "file_type, summary, key_points, extracted_fields, extracted_items, raw_text, warnings, truncated。"
        f"文件名：{file_name}；文件类型：{file_type}；是否截断：{truncated}。"
    )
    parsed = await _invoke_qwen(
        session,
        attachment=attachment,
        call_type="attachment_visual_parse",
        visual=True,
        input_payload={"file_type": file_type, "visual_count": len(urls), "truncated": truncated},
        invoke=lambda: invoke_structured(
            task=LlmTask.ATTACHMENT_VISUAL_PARSE,
            messages=[
                {
                    "role": "user",
                    "content": [
                        *[
                            {"type": "image_url", "image_url": {"url": url}}
                            for url in urls
                        ],
                        {"type": "text", "text": prompt},
                    ],
                }
            ],
            response_model=AttachmentParseJson,
            temperature=0.1,
        ),
    )
    parsed.file_type = parsed.file_type or file_type
    parsed.warnings = [*warnings, *(parsed.warnings or [])]
    parsed.truncated = bool(parsed.truncated or truncated)
    return parsed


async def _visual_url_for_attachment(
    session: AsyncSession,
    attachment: EmailAttachment,
) -> str:
    object_id = attachment.oss_object_id
    if object_id is None:
        raise ValueError("ATTACHMENT_NOT_ARCHIVED")
    return await generate_presigned_url_for_object(session, oss_object_id=object_id, expires_seconds=1800)


def _mark_attachment(
    attachment: EmailAttachment,
    *,
    status: str,
    parsed: AttachmentParseJson | None,
    text: str | None,
    error: str | None = None,
    extracted_json: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    attachment.parse_status = status
    attachment.extracted_text = text
    attachment.extracted_json = extracted_json or ({
        **parsed.model_dump(),
        "parsed_at": utcnow().isoformat(),
    } if parsed else None)
    attachment.parse_error = error
    return attachment.extracted_json


async def _parse_attachment_impl(session: AsyncSession, attachment: EmailAttachment) -> dict[str, Any] | None:
    archive_format, warnings = detect_archive_format(
        file_name=attachment.file_name,
        content_type=attachment.content_type,
    )
    if archive_format:
        attachment.content_type = ARCHIVE_CONTENT_TYPES[archive_format]
        return _mark_attachment(
            attachment,
            status="skipped",
            parsed=None,
            text=None,
            extracted_json=engineering_reference_metadata(archive_format, warnings),
        )

    file_type = attachment_type(attachment)
    if file_type not in SUPPORTED_ATTACHMENT_TYPES:
        content = None
        if attachment.oss_object_id:
            try:
                content = await download_oss_object_bytes(session, oss_object_id=attachment.oss_object_id)
            except Exception:
                content = None
        archive_format, warnings = detect_archive_format(
            file_name=attachment.file_name,
            content_type=attachment.content_type,
            content=content,
        )
        if archive_format:
            attachment.content_type = ARCHIVE_CONTENT_TYPES[archive_format]
            return _mark_attachment(
                attachment,
                status="skipped",
                parsed=None,
                text=None,
                extracted_json=engineering_reference_metadata(archive_format, warnings),
            )
        return _mark_attachment(
            attachment,
            status="unsupported",
            parsed=None,
            text=None,
            error="UNSUPPORTED_FILE_TYPE",
        )

    if not attachment.oss_object_id:
        return _mark_attachment(
            attachment,
            status="needs_manual_review",
            parsed=_fallback_json(file_type, "", warnings=["ATTACHMENT_NOT_ARCHIVED"], truncated=False),
            text=None,
            error="ATTACHMENT_NOT_ARCHIVED",
        )

    if attachment.file_size and attachment.file_size > settings.ATTACHMENT_MAX_AUTO_PARSE_BYTES:
        return _mark_attachment(
            attachment,
            status="needs_manual_review",
            parsed=_fallback_json(file_type, "", warnings=["FILE_TOO_LARGE_FOR_AUTO_PARSE"], truncated=True),
            text=None,
            error="FILE_TOO_LARGE_FOR_AUTO_PARSE",
        )

    warnings: list[str] = []
    truncated = False
    text = ""
    try:
        if file_type == "image":
            content = await download_oss_object_bytes(session, oss_object_id=attachment.oss_object_id)
            dimensions = _image_dimensions(content)
            if attachment.is_inline and dimensions and (
                dimensions[0] < settings.INLINE_IMAGE_MIN_PARSE_WIDTH
                or dimensions[1] < settings.INLINE_IMAGE_MIN_PARSE_HEIGHT
            ):
                parsed = _fallback_json(
                    file_type,
                    "",
                    warnings=["INLINE_DECORATIVE_SKIPPED"],
                    truncated=False,
                )
                parsed.extracted_fields = {"image_width": dimensions[0], "image_height": dimensions[1]}
                return _mark_attachment(
                    attachment,
                    status="skipped_decorative",
                    parsed=parsed,
                    text=None,
                )
            url = await _visual_url_for_attachment(session, attachment)
            parsed = await _qwen_visual_parse(session, attachment=attachment, file_type=file_type, file_name=attachment.file_name, urls=[url], warnings=warnings, truncated=False)
            return _mark_attachment(attachment, status="parsed", parsed=parsed, text=parsed.raw_text)

        content = await download_oss_object_bytes(session, oss_object_id=attachment.oss_object_id)
        if file_type == "txt":
            text = await _file_io(_extract_txt, content)
        elif file_type == "prc":
            text = await _file_io(_extract_prc, content)
        elif file_type == "csv":
            text = await _file_io(_extract_csv, content)
        elif file_type == "html":
            text = await _file_io(_extract_html, content)
        elif file_type == "docx":
            text = await _file_io(_extract_docx, content)
        elif file_type == "xlsx":
            text = await _file_io(_extract_xlsx, content)
        elif file_type == "pdf":
            text, page_count = await _file_io(_extract_pdf_text, content, max_pages=settings.PDF_MAX_PARSE_PAGES)
            if page_count and page_count > settings.PDF_MAX_PARSE_PAGES:
                truncated = True
                warnings.append(f"PDF_TRUNCATED_TO_{settings.PDF_MAX_PARSE_PAGES}_PAGES")
            if not text.strip():
                rendered_pages, rendered_page_count = await _file_io(
                    render_pdf_pages,
                    content,
                    max_pages=settings.PDF_MAX_PARSE_PAGES,
                )
                if rendered_page_count > settings.PDF_MAX_PARSE_PAGES and not truncated:
                    truncated = True
                    warnings.append(f"PDF_TRUNCATED_TO_{settings.PDF_MAX_PARSE_PAGES}_PAGES")
                if not rendered_pages:
                    raise ValueError("PDF_RENDER_EMPTY")
                urls = [
                    f"data:image/png;base64,{base64.b64encode(page).decode('ascii')}"
                    for page in rendered_pages
                ]
                parsed = await _qwen_visual_parse(session, attachment=attachment, file_type=file_type, file_name=attachment.file_name, urls=urls, warnings=warnings, truncated=truncated)
                return _mark_attachment(attachment, status="parsed", parsed=parsed, text=parsed.raw_text)

        text, text_truncated = _truncate_text(text)
        truncated = truncated or text_truncated
        if text_truncated:
            warnings.append("TEXT_TRUNCATED_FOR_MODEL_INPUT")
        if file_type == "txt" and text_truncated:
            parsed = _fallback_json(
                file_type,
                text,
                warnings=[*warnings, "QWEN_SKIPPED_LARGE_TEXT_LOCAL_FALLBACK"],
                truncated=True,
            )
            return _mark_attachment(attachment, status="parsed", parsed=parsed, text=text)
        parsed = await _qwen_text_parse(session, attachment=attachment, file_type=file_type, file_name=attachment.file_name, text=text, warnings=warnings, truncated=truncated)
        return _mark_attachment(attachment, status="parsed", parsed=parsed, text=parsed.raw_text or text)
    except AiProviderError as exc:
        logger.exception(
            "Attachment AI parse failed; applying fallback",
            extra={
                "event": "attachment_parse_failed", "attachment_id": attachment.id,
                "email_id": attachment.email_id, "file_type": file_type,
                "error_code": safe_error_code(exc, "ATTACHMENT_AI_FAILED"),
            },
        )
        parsed = _fallback_json(file_type, text, warnings=[*warnings, str(exc)], truncated=truncated)
        if file_type in {"txt", "prc"} and text.strip():
            parsed.warnings.append("QWEN_FAILED_LOCAL_TEXT_FALLBACK")
            return _mark_attachment(attachment, status="parsed", parsed=parsed, text=text)
        return _mark_attachment(attachment, status="needs_manual_review", parsed=parsed, text=text or None, error=str(exc))
    except (StorageConfigurationError, StorageUploadError) as exc:
        error_code = safe_error_code(exc, "ATTACHMENT_STORAGE_UNAVAILABLE")
        logger.exception(
            "Attachment storage unavailable; sending to manual review",
            extra={
                "event": "attachment_parse_failed", "attachment_id": attachment.id,
                "email_id": attachment.email_id, "file_type": file_type, "error_code": error_code,
            },
        )
        parsed = _fallback_json(file_type, text, warnings=[*warnings, error_code], truncated=truncated)
        return _mark_attachment(
            attachment, status="needs_manual_review", parsed=parsed, text=text or None,
            error=f"INFRASTRUCTURE:{error_code}",
        )
    except ValueError as exc:
        error_code = safe_error_code(exc, "ATTACHMENT_FILE_PARSE_FAILED")
        logger.exception(
            "Attachment content parse failed; sending to manual review",
            extra={
                "event": "attachment_parse_failed", "attachment_id": attachment.id,
                "email_id": attachment.email_id, "file_type": file_type, "error_code": error_code,
            },
        )
        parsed = _fallback_json(file_type, text, warnings=[*warnings, error_code], truncated=truncated)
        return _mark_attachment(
            attachment, status="needs_manual_review", parsed=parsed, text=text or None,
            error=f"FILE_CONTENT:{error_code}",
        )
    except Exception as exc:
        error_code = safe_error_code(exc, "ATTACHMENT_INFRASTRUCTURE_ERROR")
        logger.exception(
            "Attachment infrastructure failure; sending to manual review",
            extra={
                "event": "attachment_parse_failed", "attachment_id": attachment.id,
                "email_id": attachment.email_id, "file_type": file_type, "error_code": error_code,
            },
        )
        parsed = _fallback_json(file_type, text, warnings=[*warnings, error_code], truncated=truncated)
        return _mark_attachment(
            attachment, status="needs_manual_review", parsed=parsed, text=text or None,
            error=f"INFRASTRUCTURE:{error_code}",
        )


async def parse_attachment(session: AsyncSession, attachment: EmailAttachment) -> dict[str, Any] | None:
    started = time.monotonic()
    logger.info(
        "Attachment parse started",
        extra={
            "event": "attachment_parse_started", "attachment_id": attachment.id,
            "email_id": attachment.email_id, "file_type": attachment_type(attachment),
            "object_size": attachment.file_size,
        },
    )
    try:
        result = await _parse_attachment_impl(session, attachment)
    except Exception as exc:
        logger.exception(
            "Attachment parse failed",
            extra={
                "event": "attachment_parse_failed", "attachment_id": attachment.id,
                "email_id": attachment.email_id, "file_type": attachment_type(attachment),
                "duration_ms": int((time.monotonic() - started) * 1000),
                "error_code": safe_error_code(exc, "ATTACHMENT_PARSE_FAILED"),
            },
        )
        raise
    logger.info(
        "Attachment parse completed",
        extra={
            "event": "attachment_parse_completed", "attachment_id": attachment.id,
            "email_id": attachment.email_id, "file_type": attachment_type(attachment),
            "status": attachment.parse_status,
            "duration_ms": int((time.monotonic() - started) * 1000),
        },
    )
    return result
