from __future__ import annotations

import asyncio
import csv
import hashlib
import html
import io
import re
import zipfile
from datetime import date
from typing import Any
from xml.etree import ElementTree

from fastapi import HTTPException, status
from pydantic import ValidationError
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import BoardCard, JobRunLog, SnAsset
from app.core.repair_items import normalize_board_code, normalize_board_name
from app.schemas.business import BoardCardImportItem, SnAssetImportItem
from app.services.audit import log_operation
from app.services.common import model_to_dict, paginate_scalars, utcnow

SN_ASSET_FIELDS = (
    "id",
    "customer_code",
    "customer_name",
    "material_code",
    "material_name",
    "sn",
    "service_tracking_card_no",
    "parent_sn",
    "top_sn",
    "parent_material_code",
    "top_material_code",
    "asset_status",
    "warranty_start_date",
    "warranty_end_date",
    "source_file_name",
    "source_file_hash",
    "source_row_no",
    "raw_data",
    "imported_by_user_id",
    "imported_at",
    "created_at",
    "updated_at",
)

BOARD_CARD_FIELDS = (
    "id",
    "board_code",
    "board_name",
    "return_location",
    "route_type",
    "customer_scope",
    "need_ship_to_beijing",
    "shipping_address",
    "shipping_contact",
    "shipping_phone",
    "postal_code",
    "status",
    "source_file_name",
    "source_file_hash",
    "source_row_no",
    "raw_data",
    "imported_by_user_id",
    "imported_at",
    "created_at",
    "updated_at",
)

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def list_sn_assets(
    session: AsyncSession,
    *,
    page: int = 1,
    page_size: int = 20,
    keyword: str | None = None,
    sn: str | None = None,
    customer: str | None = None,
    material: str | None = None,
    asset_status: str | None = None,
) -> tuple[list[dict[str, Any]], int]:
    statement = _sn_asset_statement(keyword=keyword, sn=sn, customer=customer, material=material, asset_status=asset_status)
    statement = statement.order_by(SnAsset.updated_at.desc(), SnAsset.id.desc())
    rows, total = await paginate_scalars(session, statement, page, page_size)
    return [model_to_dict(row, SN_ASSET_FIELDS) for row in rows], total


def _sn_asset_statement(
    *,
    keyword: str | None = None,
    sn: str | None = None,
    customer: str | None = None,
    material: str | None = None,
    asset_status: str | None = None,
):
    statement = select(SnAsset)
    if asset_status:
        statement = statement.where(SnAsset.asset_status == asset_status)
    if sn:
        statement = statement.where(SnAsset.sn.like(f"%{sn.strip().upper()}%"))
    if customer:
        like = f"%{customer}%"
        statement = statement.where(or_(SnAsset.customer_code.like(like), SnAsset.customer_name.like(like)))
    if material:
        like = f"%{material}%"
        statement = statement.where(or_(SnAsset.material_code.like(like), SnAsset.material_name.like(like)))
    if keyword:
        like = f"%{keyword}%"
        statement = statement.where(
            or_(
                SnAsset.sn.like(like),
                SnAsset.service_tracking_card_no.like(like),
                SnAsset.parent_sn.like(like),
                SnAsset.top_sn.like(like),
                SnAsset.customer_code.like(like),
                SnAsset.customer_name.like(like),
                SnAsset.material_code.like(like),
                SnAsset.parent_material_code.like(like),
                SnAsset.top_material_code.like(like),
            )
        )
    return statement


async def import_sn_assets(
    session: AsyncSession,
    *,
    items: list[SnAssetImportItem],
    source_file_name: str | None,
    source_file_hash: str | None,
    user_id: int,
    job: JobRunLog | None = None,
) -> dict[str, Any]:
    owns_job = job is None
    if job is None:
        job = JobRunLog(job_name="sn_assets_import", job_type="master_data_import", status="running", processed_count=len(items), metadata_json={})
        session.add(job)
        await session.flush()
    created = 0
    updated = 0
    for item in items:
        data = item.model_dump()
        sn = data["sn"].strip().upper()
        for hierarchy_sn_field in ("parent_sn", "top_sn"):
            if data.get(hierarchy_sn_field):
                data[hierarchy_sn_field] = data[hierarchy_sn_field].strip().upper()
        row = await session.scalar(select(SnAsset).where(SnAsset.sn == sn))
        payload = {
            **data,
            "sn": sn,
            "source_file_name": source_file_name,
            "source_file_hash": source_file_hash,
            "imported_by_user_id": user_id,
            "imported_at": utcnow(),
        }
        if row is None:
            session.add(SnAsset(**payload))
            created += 1
        else:
            for key, value in payload.items():
                setattr(row, key, value)
            updated += 1
    if owns_job:
        job.status = "success"
        job.finished_at = utcnow()
    job.success_count = created + updated
    job.metadata_json = {"created": created, "updated": updated, "source_file_hash": source_file_hash}
    await log_operation(
        session,
        user_id=user_id,
        operation_type="sn_assets_imported",
        target_type="sn_assets",
        target_id=None,
        after_data=job.metadata_json,
    )
    return {"job_run_id": job.id, "created": created, "updated": updated, "processed": len(items)}


async def list_board_cards(
    session: AsyncSession,
    *,
    page: int = 1,
    page_size: int = 20,
    keyword: str | None = None,
    board_code: str | None = None,
    board_name: str | None = None,
    customer_scope: str | None = None,
    return_location: str | None = None,
    status: str | None = None,
) -> tuple[list[dict[str, Any]], int]:
    statement = _board_card_statement(
        keyword=keyword,
        board_code=board_code,
        board_name=board_name,
        customer_scope=customer_scope,
        return_location=return_location,
        status=status,
    )
    statement = statement.order_by(BoardCard.updated_at.desc(), BoardCard.id.desc())
    rows, total = await paginate_scalars(session, statement, page, page_size)
    return [model_to_dict(row, BOARD_CARD_FIELDS) for row in rows], total


def _board_card_statement(
    *,
    keyword: str | None = None,
    board_code: str | None = None,
    board_name: str | None = None,
    customer_scope: str | None = None,
    return_location: str | None = None,
    status: str | None = None,
):
    statement = select(BoardCard)
    if status:
        statement = statement.where(BoardCard.status == status)
    if board_code:
        statement = statement.where(BoardCard.board_code.like(f"%{board_code}%"))
    if board_name:
        statement = statement.where(BoardCard.board_name.like(f"%{board_name}%"))
    if customer_scope:
        statement = statement.where(BoardCard.customer_scope == customer_scope)
    if return_location:
        statement = statement.where(BoardCard.return_location == return_location)
    if keyword:
        like = f"%{keyword}%"
        statement = statement.where(
            or_(BoardCard.board_code.like(like), BoardCard.board_name.like(like))
        )
    return statement


async def import_board_cards(
    session: AsyncSession,
    *,
    items: list[BoardCardImportItem],
    source_file_name: str | None,
    source_file_hash: str | None,
    user_id: int,
    job: JobRunLog | None = None,
) -> dict[str, Any]:
    owns_job = job is None
    if job is None:
        job = JobRunLog(job_name="board_cards_import", job_type="master_data_import", status="running", processed_count=len(items), metadata_json={})
        session.add(job)
        await session.flush()
    created = 0
    updated = 0
    skipped = 0
    normalized_items: list[dict[str, Any]] = []
    active_routes: dict[str, set[str]] = {}
    for item in items:
        data = item.model_dump()
        board_code = normalize_board_code(
            data.get("board_code") or data.get("material_code")
        )
        board_name = normalize_board_name(
            data.get("board_name") or data.get("material_name")
        ) or None
        if not board_code:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="BOARD_CODE_REQUIRED",
            )
        return_location = data.get("return_location")
        if return_location is None:
            return_location = "beijing" if data.get("need_ship_to_beijing") else "tianjin"
        route_type = str(data.get("route_type") or "board_rule")
        customer_scope = str(data.get("customer_scope") or "domestic")
        normalized = {
            **data,
            "board_code": board_code,
            "board_name": board_name,
            "return_location": return_location,
            "route_type": route_type,
            "customer_scope": customer_scope,
            "material_code": board_code,
            "material_name": board_name,
            "need_ship_to_beijing": return_location == "beijing",
        }
        normalized_items.append(normalized)
        if (
            normalized.get("status", "active") == "active"
            and customer_scope == "domestic"
            and route_type == "board_rule"
        ):
            active_routes.setdefault(board_code, set()).add(return_location)
    conflicts = sorted(code for code, locations in active_routes.items() if len(locations) > 1)
    if conflicts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "BOARD_ROUTE_CONFLICT", "board_codes": conflicts},
        )

    seen: set[tuple[str, str | None, str, str]] = set()
    for data in normalized_items:
        business_key = (
            data["board_code"],
            data["board_name"],
            data["customer_scope"],
            data["route_type"],
        )
        if business_key in seen:
            skipped += 1
            continue
        seen.add(business_key)
        if (
            data["customer_scope"] == "overseas"
            and data["route_type"] == "scope_default"
            and data.get("status", "active") == "active"
        ):
            duplicate_default = await session.scalar(
                select(BoardCard).where(
                    BoardCard.customer_scope == "overseas",
                    BoardCard.route_type == "scope_default",
                    BoardCard.status == "active",
                    or_(
                        BoardCard.board_code != data["board_code"],
                        BoardCard.board_name != data["board_name"],
                    ),
                )
            )
            if duplicate_default is not None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="OVERSEAS_DEFAULT_ROUTE_ALREADY_EXISTS",
                )
        row = await session.scalar(
            select(BoardCard).where(
                BoardCard.board_code == data["board_code"],
                BoardCard.board_name == data["board_name"],
                BoardCard.customer_scope == data["customer_scope"],
                BoardCard.route_type == data["route_type"],
            )
        )
        payload = {
            **data,
            "source_file_name": source_file_name,
            "source_file_hash": source_file_hash,
            "imported_by_user_id": user_id,
            "imported_at": utcnow(),
        }
        if row is None:
            session.add(BoardCard(**payload))
            created += 1
        else:
            for key, value in payload.items():
                setattr(row, key, value)
            updated += 1
    if owns_job:
        job.status = "success"
        job.finished_at = utcnow()
    job.success_count = created + updated
    job.metadata_json = {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "source_file_hash": source_file_hash,
    }
    await log_operation(
        session,
        user_id=user_id,
        operation_type="board_cards_imported",
        target_type="board_cards",
        target_id=None,
        after_data=job.metadata_json,
    )
    return {
        "job_run_id": job.id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "processed": len(items),
    }


def _date_or_none(value: str | None) -> date | None:
    value = (value or "").strip()
    return date.fromisoformat(value) if value else None


def _bool_value(value: str | None) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in {"1", "true", "yes", "y", "是", "启用"}


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _xlsx_with_openpyxl(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes | None:
    try:
        from openpyxl import Workbook  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([_string_value(row.get(field)) for field in fieldnames])
    for column_cells in sheet.columns:
        max_length = max(len(_string_value(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheet_title(value: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]\:\*\?\/\\]", "_", value).strip() or "sheet"
    title = title[:31]
    candidate = title
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _xlsx_workbook_with_openpyxl(sheets: list[tuple[str, list[dict[str, Any]], list[str]]]) -> bytes | None:
    try:
        from openpyxl import Workbook  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()
    for title, rows, fieldnames in sheets:
        sheet = workbook.create_sheet(_sheet_title(title, used))
        sheet.append(fieldnames)
        for row in rows:
            sheet.append([_string_value(row.get(field)) for field in fieldnames])
        for column_cells in sheet.columns:
            max_length = max(len(_string_value(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _xlsx_fallback(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes:
    def cell(row_index: int, column_index: int, value: Any) -> str:
        text = html.escape(_string_value(value), quote=False)
        ref = f"{_column_name(column_index)}{row_index}"
        return f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'

    sheet_rows: list[str] = []
    all_rows = [dict(zip(fieldnames, fieldnames))] + rows
    for row_index, row in enumerate(all_rows, start=1):
        cells = "".join(cell(row_index, column_index, row.get(field)) for column_index, field in enumerate(fieldnames, start=1))
        sheet_rows.append(f'<row r="{row_index}">{cells}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="data" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/></Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buffer.getvalue()


def _xlsx_workbook_fallback(sheets: list[tuple[str, list[dict[str, Any]], list[str]]]) -> bytes:
    worksheets: list[str] = []
    sheet_entries: list[str] = []
    relationship_entries: list[str] = []
    content_type_entries: list[str] = []
    used: set[str] = set()
    for sheet_index, (title, rows, fieldnames) in enumerate(sheets, start=1):
        def cell(row_index: int, column_index: int, value: Any) -> str:
            text = html.escape(_string_value(value), quote=False)
            ref = f"{_column_name(column_index)}{row_index}"
            return f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'

        sheet_rows: list[str] = []
        all_rows = [dict(zip(fieldnames, fieldnames))] + rows
        for row_index, row in enumerate(all_rows, start=1):
            cells = "".join(cell(row_index, column_index, row.get(field)) for column_index, field in enumerate(fieldnames, start=1))
            sheet_rows.append(f'<row r="{row_index}">{cells}</row>')
        worksheets.append(
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f"<sheetData>{''.join(sheet_rows)}</sheetData>"
            "</worksheet>"
        )
        sheet_name = html.escape(_sheet_title(title, used), quote=True)
        sheet_entries.append(f'<sheet name="{sheet_name}" sheetId="{sheet_index}" r:id="rId{sheet_index}"/>')
        relationship_entries.append(
            f'<Relationship Id="rId{sheet_index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{sheet_index}.xml"/>'
        )
        content_type_entries.append(
            f'<Override PartName="/xl/worksheets/sheet{sheet_index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{''.join(sheet_entries)}</sheets></workbook>"
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{''.join(relationship_entries)}</Relationships>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f"{''.join(content_type_entries)}</Types>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        for index, worksheet in enumerate(worksheets, start=1):
            archive.writestr(f"xl/worksheets/sheet{index}.xml", worksheet)
    return buffer.getvalue()


def xlsx_bytes(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes:
    return _xlsx_with_openpyxl(rows, fieldnames) or _xlsx_fallback(rows, fieldnames)


def xlsx_workbook_bytes(sheets: list[tuple[str, list[dict[str, Any]], list[str]]]) -> bytes:
    if not sheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="EXPORT_SELECTION_REQUIRED")
    return _xlsx_workbook_with_openpyxl(sheets) or _xlsx_workbook_fallback(sheets)


def _read_xlsx_with_openpyxl(content: bytes) -> list[dict[str, Any]] | None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX_HEADER_REQUIRED")
    headers = [_string_value(value).strip() for value in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX_HEADER_REQUIRED")
    data: list[dict[str, Any]] = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        data.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]})
    return data


def _cell_index(ref: str | None) -> int:
    if not ref:
        return 0
    match = re.match(r"([A-Z]+)", ref.upper())
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - 64
    return index - 1


def _read_xlsx_fallback(content: bytes) -> list[dict[str, Any]]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX_INVALID_FILE") from exc

    shared_strings: list[str] = []
    try:
        shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
        for item in shared_root.findall(".//{*}si"):
            shared_strings.append("".join(text.text or "" for text in item.findall(".//{*}t")))
    except KeyError:
        pass

    root = ElementTree.fromstring(sheet_xml)
    parsed_rows: list[list[str]] = []
    for row in root.findall(".//{*}row"):
        values: list[str] = []
        for cell in row.findall("{*}c"):
            column_index = _cell_index(cell.attrib.get("r"))
            while len(values) <= column_index:
                values.append("")
            cell_type = cell.attrib.get("t")
            if cell_type == "inlineStr":
                value = "".join(text.text or "" for text in cell.findall(".//{*}t"))
            else:
                raw = cell.find("{*}v")
                value = raw.text if raw is not None and raw.text is not None else ""
                if cell_type == "s" and value.isdigit() and int(value) < len(shared_strings):
                    value = shared_strings[int(value)]
            values[column_index] = value
        if any(value != "" for value in values):
            parsed_rows.append(values)
    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX_HEADER_REQUIRED")
    headers = [value.strip() for value in parsed_rows[0]]
    if not any(headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX_HEADER_REQUIRED")
    return [
        {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers)) if headers[index]}
        for row in parsed_rows[1:]
    ]


def _read_xlsx(content: bytes) -> tuple[list[dict[str, Any]], str]:
    rows = _read_xlsx_with_openpyxl(content)
    if rows is None:
        rows = _read_xlsx_fallback(content)
    return rows, hashlib.sha256(content).hexdigest()


def parse_sn_assets_xlsx(content: bytes) -> tuple[list[SnAssetImportItem], str]:
    rows, file_hash = _read_xlsx(content)
    items: list[SnAssetImportItem] = []
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        try:
            items.append(
                SnAssetImportItem(
                    customer_code=_string_value(row.get("customer_code")).strip(),
                    customer_name=_string_value(row.get("customer_name")).strip(),
                    material_code=_string_value(row.get("material_code")).strip(),
                    material_name=_string_value(row.get("material_name")).strip() or None,
                    service_tracking_card_no=_string_value(row.get("service_tracking_card_no")).strip() or None,
                    parent_sn=_string_value(row.get("parent_sn")).strip().upper() or None,
                    top_sn=_string_value(row.get("top_sn")).strip().upper() or None,
                    parent_material_code=_string_value(row.get("parent_material_code")).strip() or None,
                    top_material_code=_string_value(row.get("top_material_code")).strip() or None,
                    sn=_string_value(row.get("sn")).strip(),
                    asset_status=_string_value(row.get("asset_status") or "valid").strip(),
                    warranty_start_date=row.get("warranty_start_date")
                    if isinstance(row.get("warranty_start_date"), date)
                    else _date_or_none(_string_value(row.get("warranty_start_date"))),
                    warranty_end_date=row.get("warranty_end_date")
                    if isinstance(row.get("warranty_end_date"), date)
                    else _date_or_none(_string_value(row.get("warranty_end_date"))),
                    source_row_no=index,
                    raw_data={key: _string_value(value) for key, value in row.items()},
                )
            )
        except (ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc)})
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"code": "XLSX_VALIDATION_FAILED", "errors": errors})
    return items, file_hash


def parse_board_cards_xlsx(content: bytes) -> tuple[list[BoardCardImportItem], str]:
    rows, file_hash = _read_xlsx(content)
    items: list[BoardCardImportItem] = []
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        try:
            items.append(
                BoardCardImportItem(
                    board_code=_string_value(
                        row.get("board_code") or row.get("material_code")
                    ).strip(),
                    board_name=_string_value(
                        row.get("board_name") or row.get("material_name")
                    ).strip() or None,
                    return_location=_string_value(row.get("return_location")).strip() or None,
                    route_type=_string_value(row.get("route_type") or "board_rule").strip(),
                    customer_scope=_string_value(
                        row.get("customer_scope") or "domestic"
                    ).strip(),
                    material_code=_string_value(row.get("material_code")).strip() or None,
                    material_name=_string_value(row.get("material_name")).strip() or None,
                    need_ship_to_beijing=(
                        _bool_value(_string_value(row.get("need_ship_to_beijing")))
                        if row.get("need_ship_to_beijing") is not None
                        else None
                    ),
                    shipping_address=_string_value(row.get("shipping_address")).strip() or None,
                    shipping_contact=_string_value(row.get("shipping_contact")).strip() or None,
                    shipping_phone=_string_value(row.get("shipping_phone")).strip() or None,
                    postal_code=_string_value(row.get("postal_code")).strip() or None,
                    status=_string_value(row.get("status") or "active").strip(),
                    source_row_no=index,
                    raw_data={key: _string_value(value) for key, value in row.items()},
                )
            )
        except (ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc)})
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"code": "XLSX_VALIDATION_FAILED", "errors": errors})
    return items, file_hash


def _split_return_address(value: str) -> tuple[str | None, str | None, str | None, str | None]:
    normalized = re.sub(r"\s+", " ", value or "").strip()
    if not normalized:
        return None, None, None, None
    phone_match = re.search(r"(?:010|022)[-\d]+(?:-\d+)?", normalized)
    postal_match = re.search(r"邮编\s*[:：]?\s*(\d{6})", normalized)
    contact_match = re.search(
        r"([\u4e00-\u9fff]{2,10})(?:（收）|\(收\))?\s*(?:电话|TEL)",
        normalized,
        flags=re.IGNORECASE,
    )
    contact = contact_match.group(1) if contact_match else None
    address = normalized
    cut_positions = [
        match.start()
        for match in (contact_match, phone_match, postal_match)
        if match is not None
    ]
    if cut_positions:
        address = normalized[: min(cut_positions)].strip(" ,，;；")
    return (
        address or None,
        contact,
        phone_match.group(0) if phone_match else None,
        postal_match.group(1) if postal_match else None,
    )


def parse_board_cards_xls(content: bytes) -> tuple[list[BoardCardImportItem], str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="XLRD_NOT_INSTALLED",
        ) from exc

    workbook = xlrd.open_workbook(file_contents=content)
    items: list[BoardCardImportItem] = []
    for sheet_index, sheet in enumerate(workbook.sheets()):
        current_raw_address = ""
        location = "beijing" if sheet_index == 0 else "tianjin"
        for row_index in range(2, sheet.nrows):
            board_code = _string_value(sheet.cell_value(row_index, 0)).strip().upper()
            board_name = _string_value(sheet.cell_value(row_index, 1)).strip() or None
            raw_address = _string_value(sheet.cell_value(row_index, 2)).strip()
            if raw_address:
                current_raw_address = raw_address
                if "北京" in raw_address:
                    location = "beijing"
                elif "天津" in raw_address:
                    location = "tianjin"
            if not board_code:
                continue
            address, contact, phone, postal_code = _split_return_address(
                current_raw_address
            )
            items.append(
                BoardCardImportItem(
                    board_code=board_code,
                    board_name=board_name,
                    return_location=location,
                    route_type="board_rule",
                    customer_scope="domestic",
                    shipping_address=address,
                    shipping_contact=contact,
                    shipping_phone=phone,
                    postal_code=postal_code,
                    status="active",
                    source_row_no=row_index + 1,
                    raw_data={
                        "sheet_name": sheet.name,
                        "raw_shipping_address": current_raw_address,
                    },
                )
            )
    return items, hashlib.sha256(content).hexdigest()


def parse_board_cards_file(
    content: bytes,
    *,
    filename: str | None,
) -> tuple[list[BoardCardImportItem], str]:
    if (filename or "").lower().endswith(".xls"):
        return parse_board_cards_xls(content)
    return parse_board_cards_xlsx(content)


def _read_csv(content: bytes) -> tuple[list[dict[str, str]], str]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV_HEADER_REQUIRED")
    return [dict(row) for row in reader], hashlib.sha256(content).hexdigest()


def parse_sn_assets_csv(content: bytes) -> tuple[list[SnAssetImportItem], str]:
    rows, file_hash = _read_csv(content)
    items: list[SnAssetImportItem] = []
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        try:
            items.append(
                SnAssetImportItem(
                    customer_code=(row.get("customer_code") or "").strip(),
                    customer_name=(row.get("customer_name") or "").strip(),
                    material_code=(row.get("material_code") or "").strip(),
                    material_name=(row.get("material_name") or "").strip() or None,
                    service_tracking_card_no=(row.get("service_tracking_card_no") or "").strip() or None,
                    parent_sn=(row.get("parent_sn") or "").strip().upper() or None,
                    top_sn=(row.get("top_sn") or "").strip().upper() or None,
                    parent_material_code=(row.get("parent_material_code") or "").strip() or None,
                    top_material_code=(row.get("top_material_code") or "").strip() or None,
                    sn=(row.get("sn") or "").strip(),
                    asset_status=(row.get("asset_status") or "valid").strip(),
                    warranty_start_date=_date_or_none(row.get("warranty_start_date")),
                    warranty_end_date=_date_or_none(row.get("warranty_end_date")),
                    source_row_no=index,
                    raw_data=row,
                )
            )
        except (ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc)})
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"code": "CSV_VALIDATION_FAILED", "errors": errors})
    return items, file_hash


def parse_board_cards_csv(content: bytes) -> tuple[list[BoardCardImportItem], str]:
    rows, file_hash = _read_csv(content)
    items: list[BoardCardImportItem] = []
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        try:
            items.append(
                BoardCardImportItem(
                    board_code=(row.get("board_code") or row.get("material_code") or "").strip(),
                    board_name=(row.get("board_name") or row.get("material_name") or "").strip() or None,
                    return_location=(row.get("return_location") or "").strip() or None,
                    route_type=(row.get("route_type") or "board_rule").strip(),
                    customer_scope=(row.get("customer_scope") or "domestic").strip(),
                    material_code=(row.get("material_code") or "").strip() or None,
                    material_name=(row.get("material_name") or "").strip() or None,
                    need_ship_to_beijing=(
                        _bool_value(row.get("need_ship_to_beijing"))
                        if row.get("need_ship_to_beijing") not in {None, ""}
                        else None
                    ),
                    shipping_address=(row.get("shipping_address") or "").strip() or None,
                    shipping_contact=(row.get("shipping_contact") or "").strip() or None,
                    shipping_phone=(row.get("shipping_phone") or "").strip() or None,
                    postal_code=(row.get("postal_code") or "").strip() or None,
                    status=(row.get("status") or "active").strip(),
                    source_row_no=index,
                    raw_data=row,
                )
            )
        except (ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc)})
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"code": "CSV_VALIDATION_FAILED", "errors": errors})
    return items, file_hash


def csv_bytes(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def sn_assets_template_csv() -> bytes:
    return csv_bytes(
        [
            {
                "sn": "SN202607070001",
                "customer_code": "CUST001",
                "customer_name": "示例客户",
                "material_code": "MAT001",
                "material_name": "示例物料",
                "service_tracking_card_no": "STC202607070001",
                "parent_sn": "SN-PARENT-001",
                "top_sn": "SN-TOP-001",
                "parent_material_code": "MAT-PARENT-001",
                "top_material_code": "MAT-TOP-001",
                "asset_status": "valid",
                "warranty_start_date": "2026-01-01",
                "warranty_end_date": "2027-01-01",
            }
        ],
        [
            "sn", "customer_code", "customer_name", "material_code", "material_name",
            "service_tracking_card_no", "parent_sn", "top_sn", "parent_material_code",
            "top_material_code", "asset_status", "warranty_start_date", "warranty_end_date",
        ],
    )


def board_cards_template_csv() -> bytes:
    return csv_bytes(
        [
            {
                "board_code": "M8002",
                "board_name": "PVI",
                "return_location": "beijing",
                "route_type": "board_rule",
                "customer_scope": "domestic",
                "shipping_address": "北京市示例地址",
                "shipping_contact": "张三",
                "shipping_phone": "010-00000000",
                "postal_code": "100000",
                "status": "active",
            }
        ],
        [
            "board_code", "board_name", "return_location", "route_type",
            "customer_scope", "shipping_address", "shipping_contact",
            "shipping_phone", "postal_code", "status",
        ],
    )


def sn_assets_template_xlsx() -> bytes:
    return xlsx_bytes(
        [
            {
                "sn": "SN202607070001",
                "customer_code": "CUST001",
                "customer_name": "示例客户",
                "material_code": "MAT001",
                "material_name": "示例物料",
                "service_tracking_card_no": "STC202607070001",
                "parent_sn": "SN-PARENT-001",
                "top_sn": "SN-TOP-001",
                "parent_material_code": "MAT-PARENT-001",
                "top_material_code": "MAT-TOP-001",
                "asset_status": "valid",
                "warranty_start_date": "2026-01-01",
                "warranty_end_date": "2027-01-01",
            }
        ],
        [
            "sn", "customer_code", "customer_name", "material_code", "material_name",
            "service_tracking_card_no", "parent_sn", "top_sn", "parent_material_code",
            "top_material_code", "asset_status", "warranty_start_date", "warranty_end_date",
        ],
    )


def board_cards_template_xlsx() -> bytes:
    return xlsx_bytes(
        [
            {
                "board_code": "M8002",
                "board_name": "PVI",
                "return_location": "beijing",
                "route_type": "board_rule",
                "customer_scope": "domestic",
                "shipping_address": "北京市示例地址",
                "shipping_contact": "张三",
                "shipping_phone": "010-00000000",
                "postal_code": "100000",
                "status": "active",
            }
        ],
        [
            "board_code", "board_name", "return_location", "route_type",
            "customer_scope", "shipping_address", "shipping_contact",
            "shipping_phone", "postal_code", "status",
        ],
    )


async def export_sn_assets(
    session: AsyncSession,
    *,
    keyword: str | None = None,
    sn: str | None = None,
    customer: str | None = None,
    material: str | None = None,
    asset_status: str | None = None,
) -> bytes:
    statement = _sn_asset_statement(keyword=keyword, sn=sn, customer=customer, material=material, asset_status=asset_status).order_by(
        SnAsset.updated_at.desc(), SnAsset.id.desc()
    )
    rows = (await session.execute(statement)).scalars().all()
    return await asyncio.to_thread(xlsx_bytes, [model_to_dict(row, SN_ASSET_FIELDS) for row in rows], list(SN_ASSET_FIELDS))


async def export_sn_assets_selected(session: AsyncSession, *, ids: list[int]) -> bytes:
    if not ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="EXPORT_SELECTION_REQUIRED")
    rows = (await session.execute(select(SnAsset).where(SnAsset.id.in_(ids)).order_by(SnAsset.id))).scalars().all()
    return await asyncio.to_thread(xlsx_bytes, [model_to_dict(row, SN_ASSET_FIELDS) for row in rows], list(SN_ASSET_FIELDS))


async def export_board_cards(
    session: AsyncSession,
    *,
    keyword: str | None = None,
    board_code: str | None = None,
    board_name: str | None = None,
    customer_scope: str | None = None,
    return_location: str | None = None,
    status: str | None = None,
) -> bytes:
    statement = _board_card_statement(
        keyword=keyword,
        board_code=board_code,
        board_name=board_name,
        customer_scope=customer_scope,
        return_location=return_location,
        status=status,
    ).order_by(
        BoardCard.updated_at.desc(), BoardCard.id.desc()
    )
    rows = (await session.execute(statement)).scalars().all()
    return await asyncio.to_thread(xlsx_bytes, [model_to_dict(row, BOARD_CARD_FIELDS) for row in rows], list(BOARD_CARD_FIELDS))


async def export_board_cards_selected(session: AsyncSession, *, ids: list[int]) -> bytes:
    if not ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="EXPORT_SELECTION_REQUIRED")
    rows = (await session.execute(select(BoardCard).where(BoardCard.id.in_(ids)).order_by(BoardCard.id))).scalars().all()
    return await asyncio.to_thread(xlsx_bytes, [model_to_dict(row, BOARD_CARD_FIELDS) for row in rows], list(BOARD_CARD_FIELDS))
