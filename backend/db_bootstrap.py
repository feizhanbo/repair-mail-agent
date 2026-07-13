from __future__ import annotations

import asyncio
import hashlib
import os
from urllib.parse import urlparse

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine

from app.config import BACKEND_DIR, settings
from app.core.database import AsyncSessionLocal
from app.schemas.business import BoardCardImportItem, SnAssetImportItem
from app.services.master_data import import_board_cards, import_sn_assets

DB_NAME = "repair_system_test"
SN_XLSX_PATH = r"D:\refile\SNdata.xlsx"
BOARD_CARD_XLS_PATH = r"D:\refile\寄北京板卡.xls"


def _parse_url(url: str) -> tuple[str, str, str, str, int | None, str | None]:
    parsed = urlparse(url)
    return (
        parsed.scheme,
        parsed.username or "",
        parsed.password or "",
        parsed.hostname or "",
        parsed.port,
        parsed.path.lstrip("/") or None,
    )


def _build_url(scheme: str, user: str, password: str, host: str, port: int | None, db_name: str | None = None) -> str:
    auth = f"{user}:{password}" if password else user
    port_part = f":{port}" if port else ""
    db_part = f"/{db_name}" if db_name else ""
    return f"{scheme}://{auth}@{host}{port_part}{db_part}"


async def _run_subprocess(cmd: str, cwd: str, env: dict[str, str]) -> None:
    process = await asyncio.create_subprocess_shell(
        cmd,
        cwd=cwd,
        env=env,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await process.communicate()
    if stdout:
        print(stdout.decode().rstrip())
    if stderr:
        print(stderr.decode().rstrip())
    if process.returncode != 0:
        raise RuntimeError(f"command failed with exit code {process.returncode}: {cmd}")


async def _check_database_exists(nodb_url: str, db_name: str) -> bool:
    engine = create_async_engine(nodb_url)
    try:
        async with engine.connect() as conn:
            result = await conn.execute(text("SHOW DATABASES LIKE :db"), {"db": db_name})
            return result.fetchone() is not None
    finally:
        await engine.dispose()


async def _create_database(nodb_url: str, db_name: str) -> None:
    engine = create_async_engine(nodb_url)
    try:
        async with engine.connect() as conn:
            await conn.execute(
                text(
                    f"CREATE DATABASE `{db_name}` "
                    "DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                )
            )
            await conn.commit()
    finally:
        await engine.dispose()


async def _check_seed_data(db_url: str) -> bool:
    engine = create_async_engine(db_url)
    try:
        async with engine.connect() as conn:
            result = await conn.execute(text("SELECT COUNT(*) FROM workflow_statuses"))
            count = result.scalar_one_or_none()
            return count is not None and count > 0
    except Exception:
        return False
    finally:
        await engine.dispose()


async def _import_sn_assets_from_xlsx() -> dict:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        print("openpyxl not installed, skipping SN asset import.")
        return {"created": 0, "updated": 0}

    if not os.path.exists(SN_XLSX_PATH):
        print(f"SN asset file not found: {SN_XLSX_PATH}, skipping.")
        return {"created": 0, "updated": 0}

    print(f"Importing SN assets from SNdata.xlsx...")

    workbook = load_workbook(SN_XLSX_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("  SN asset file is empty, skipping.")
        workbook.close()
        return {"created": 0, "updated": 0}

    header_map = {
        "客户": "customer_code",
        "客户名称": "customer_name",
        "物料代码": "material_code",
        "物料名称": "material_name",
        "SN": "sn",
    }
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    mapped_headers = [header_map.get(h, h) for h in headers]

    items: list[SnAssetImportItem] = []
    for idx, values in enumerate(rows[1:], start=2):
        row = {mapped_headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(values) if i < len(mapped_headers)}
        if not row.get("sn"):
            continue
        items.append(
            SnAssetImportItem(
                customer_code=row.get("customer_code", ""),
                customer_name=row.get("customer_name", ""),
                material_code=row.get("material_code", ""),
                material_name=row.get("material_name") or None,
                sn=row.get("sn", ""),
                asset_status="valid",
                warranty_start_date=None,
                warranty_end_date=None,
                source_row_no=idx,
                raw_data=row,
            )
        )
    workbook.close()

    if not items:
        print("  No valid SN asset rows found.")
        return {"created": 0, "updated": 0}

    with open(SN_XLSX_PATH, "rb") as f:
        content = f.read()
    file_hash = hashlib.sha256(content).hexdigest()

    async with AsyncSessionLocal() as session:
        async with session.begin():
            result = await import_sn_assets(
                session,
                items=items,
                source_file_name="SNdata.xlsx",
                source_file_hash=file_hash,
                user_id=1,
            )

    print(f"  SN assets: {result['created']} created, {result['updated']} updated.")
    return {"created": result["created"], "updated": result["updated"]}


async def _import_board_cards_from_xls() -> dict:
    try:
        import xlrd
    except ModuleNotFoundError:
        print("xlrd not installed, skipping board card import.")
        return {"created": 0, "updated": 0}

    if not os.path.exists(BOARD_CARD_XLS_PATH):
        print(f"Board card file not found: {BOARD_CARD_XLS_PATH}, skipping.")
        return {"created": 0, "updated": 0}

    print(f"Importing board cards from 寄北京板卡.xls...")

    workbook = xlrd.open_workbook(BOARD_CARD_XLS_PATH)
    sheet = workbook.sheet_by_index(0)

    if sheet.nrows < 3:
        print("  Board card file has no data rows, skipping.")
        return {"created": 0, "updated": 0}

    header_map = {
        "板卡型号": "material_code",
        "板卡名称": "material_name",
        "收货地址": "shipping_address",
    }
    header_row = sheet.row_values(1)
    mapped_headers = [header_map.get(str(v).strip(), str(v).strip()) for v in header_row]

    items: list[BoardCardImportItem] = []
    for row_idx in range(2, sheet.nrows):
        values = sheet.row_values(row_idx)
        row = {mapped_headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(values) if i < len(mapped_headers)}
        material_code = row.get("material_code", "")
        if not material_code:
            continue
        items.append(
            BoardCardImportItem(
                material_code=material_code,
                material_name=row.get("material_name") or None,
                need_ship_to_beijing=True,
                shipping_address=row.get("shipping_address") or None,
                shipping_contact=None,
                shipping_phone=None,
                postal_code=None,
                status="active",
                source_row_no=row_idx + 1,
                raw_data=row,
            )
        )

    if not items:
        print("  No valid board card rows found.")
        return {"created": 0, "updated": 0}

    with open(BOARD_CARD_XLS_PATH, "rb") as f:
        content = f.read()
    file_hash = hashlib.sha256(content).hexdigest()

    try:
        async with AsyncSessionLocal() as session:
            async with session.begin():
                result = await import_board_cards(
                    session,
                    items=items,
                    source_file_name="寄北京板卡.xls",
                    source_file_hash=file_hash,
                    user_id=1,
                )
        print(f"  Board cards: {result['created']} created, {result['updated']} updated.")
        return {"created": result["created"], "updated": result["updated"]}
    except Exception as exc:
        print(f"  Board card import failed (table may not exist yet): {exc}")
        return {"created": 0, "updated": 0}


async def _main() -> None:
    scheme, user, password, host, port, db_name = _parse_url(settings.DATABASE_URL)
    nodb_url = _build_url(scheme, user, password, host, port)
    test_db_url = _build_url(scheme, user, password, host, port, DB_NAME)

    try:
        db_exists = await _check_database_exists(nodb_url, DB_NAME)
    except Exception as exc:
        print(f"Database connection failed: {exc}")
        print("Please verify SSH tunnel / MySQL credentials and try again.")
        raise SystemExit(1) from exc

    env = os.environ.copy()
    env["DATABASE_URL"] = test_db_url

    if not db_exists:
        print(f"Database '{DB_NAME}' does not exist, creating...")
        try:
            await _create_database(nodb_url, DB_NAME)
        except Exception as exc:
            print(f"  Failed to create database: {exc}")
            raise SystemExit(1) from exc
        print("  Database created.")

        print("Running Alembic migrations...")
        await _run_subprocess("alembic upgrade head", str(BACKEND_DIR), env)

        print("Running seed data...")
        await _run_subprocess("python -m app.seed", str(BACKEND_DIR), env)
    else:
        print(f"Database '{DB_NAME}' already exists, checking seed data...")
        has_seed = await _check_seed_data(test_db_url)
        if not has_seed:
            print("Seed data not found, running migration and seed...")
            print("  Running Alembic migrations...")
            await _run_subprocess("alembic upgrade head", str(BACKEND_DIR), env)
            print("  Running seed data...")
            await _run_subprocess("python -m app.seed", str(BACKEND_DIR), env)
        else:
            print("Seed data already present, skipping migration/seed.")

    sn_result = await _import_sn_assets_from_xlsx()
    board_result = await _import_board_cards_from_xls()

    print()
    print(
        f"Bootstrap complete. "
        f"SN: {sn_result['created']} created, {sn_result['updated']} updated. "
        f"Board cards: {board_result['created']} created, {board_result['updated']} updated."
    )


if __name__ == "__main__":
    asyncio.run(_main())
