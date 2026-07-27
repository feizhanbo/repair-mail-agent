"""
Sync SN data from xlsx test attachment into sn_assets table.

Usage: python tests/sync_test_sns.py
  (run from backend directory)
"""

from __future__ import annotations

import asyncio
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

# Ensure backend root is on sys.path so `app` is importable
BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.database import engine  # noqa: E402

XLSX_PATH = Path(r"D:\refile\emlattachment\05_15个SN_动态扩展测试.xlsx")

# ---------------------------------------------------------------------------
# Column header → sn_assets column mapping keywords
# ---------------------------------------------------------------------------
SN_KEYWORDS = [
    "板卡编号", "part serial", "sn", "序列号",
    "设备编号", "机身编号",
]
# Keywords that indicate a row-counter column, not an SN column
SN_EXCLUDE_KEYWORDS = [
    "序号",  # sequence number / row index, not serial number
]
CUSTOMER_CODE_KEYWORDS = [
    "客户代码", "customer code", "客户编码", "客户编号",
]
CUSTOMER_NAME_KEYWORDS = [
    "客户名称", "customer name", "客户", "公司名称", "公司名",
]
MATERIAL_CODE_KEYWORDS = [
    "物料代码", "material code", "物料编码", "板卡型号", "board model",
    "产品型号", "型号",
]
MATERIAL_NAME_KEYWORDS = [
    "物料名称", "material name", "物料描述", "板卡名称", "产品名称",
]
FAILURE_DESC_KEYWORDS = [
    "故障描述", "failure description", "故障现象", "故障说明",
    "故障信息", "failure information",
]


def _normalize(s: str) -> str:
    """Lower-case and strip whitespace for keyword matching."""
    if s is None:
        return ""
    return s.strip().lower()


def _match_keyword(cell_text: str, keywords: list[str]) -> bool:
    """Check if cell_text contains any of the keywords (case-insensitive)."""
    t = _normalize(cell_text)
    for kw in keywords:
        if kw in t:
            return True
    return False


def parse_xlsx(xlsx_path: Path) -> tuple[list[dict], dict]:
    """Parse the xlsx and return (rows, form_meta).

    rows: list of dicts with keys sn, customer_code, customer_name,
          material_code, material_name, failure_description
    form_meta: dict with any global form-level fields extracted from
               metadata rows (e.g. company name in an RMA form).
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    all_rows: list[list[str | None]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        all_rows.append([str(v).strip() if v is not None else None for v in row])

    # ------------------------------------------------------------------
    # 1. Locate the header row – the first row that contains column labels
    #    like SN / 板卡编号 / 物料代码 / 客户, etc.
    # ------------------------------------------------------------------
    header_row_idx: int | None = None
    header_texts: list[str] = []

    for i, row_data in enumerate(all_rows):
        combined = " ".join(_normalize(c) for c in row_data if c)
        # Check if this row looks like a header row
        has_sn = _match_keyword(combined, SN_KEYWORDS)
        has_material = _match_keyword(combined, MATERIAL_CODE_KEYWORDS)
        has_customer = _match_keyword(combined, CUSTOMER_CODE_KEYWORDS + CUSTOMER_NAME_KEYWORDS)
        has_fail = _match_keyword(combined, FAILURE_DESC_KEYWORDS)
        # At minimum we need SN and one more business column
        if has_sn and (has_material or has_customer or has_fail):
            header_row_idx = i
            header_texts = [_normalize(c) for c in row_data]
            break

    if header_row_idx is None:
        # Fallback: treat row 0 as header
        header_row_idx = 0
        header_texts = [_normalize(c) for c in all_rows[0]]

    # ------------------------------------------------------------------
    # 2. Map header columns to sn_assets fields
    # ------------------------------------------------------------------
    col_map: dict[str, int] = {}  # field_name -> col index
    for col_idx, h in enumerate(header_texts):
        if not h:
            continue
        if _match_keyword(h, SN_KEYWORDS) and not _match_keyword(h, SN_EXCLUDE_KEYWORDS):
            col_map.setdefault("sn", col_idx)
        if _match_keyword(h, CUSTOMER_CODE_KEYWORDS):
            col_map.setdefault("customer_code", col_idx)
        if _match_keyword(h, CUSTOMER_NAME_KEYWORDS):
            col_map.setdefault("customer_name", col_idx)
        if _match_keyword(h, MATERIAL_CODE_KEYWORDS):
            col_map.setdefault("material_code", col_idx)
        if _match_keyword(h, MATERIAL_NAME_KEYWORDS):
            col_map.setdefault("material_name", col_idx)
        if _match_keyword(h, FAILURE_DESC_KEYWORDS):
            col_map.setdefault("failure_description", col_idx)

    if "sn" not in col_map:
        raise ValueError(f"Could not find SN column in header row {header_row_idx + 1}: {header_texts}")

    # ------------------------------------------------------------------
    # 3. Extract form-level metadata from rows before the header
    #    (e.g. customer name / company name in RMA forms)
    # ------------------------------------------------------------------
    form_meta: dict[str, str] = {}
    for row_data in all_rows[:header_row_idx]:
        for col_idx, cell in enumerate(row_data):
            if not cell:
                continue
            c = _normalize(cell)
            if _match_keyword(c, CUSTOMER_NAME_KEYWORDS):
                # Look for the value in the next column, or this same row
                for offset in [1, -1, 2]:
                    val_col = col_idx + offset
                    if 0 <= val_col < len(row_data) and row_data[val_col]:
                        val = str(row_data[val_col]).strip()
                        if val:
                            form_meta["customer_name"] = val
                            break

    # ------------------------------------------------------------------
    # 4. Parse data rows (everything below the header)
    # ------------------------------------------------------------------
    rows: list[dict] = []
    for row_data in all_rows[header_row_idx + 1:]:
        sn_val = row_data[col_map["sn"]] if "sn" in col_map and col_map["sn"] < len(row_data) else None
        if not sn_val or len(str(sn_val).strip()) < 5:
            # Skip summary / sub-total / accessory / empty rows
            continue

        row: dict = {}
        row["sn"] = str(sn_val).strip().replace("\n", " ").replace("\r", "")

        for field in ["customer_code", "customer_name", "material_code", "material_name", "failure_description"]:
            if field in col_map and col_map[field] < len(row_data):
                val = row_data[col_map[field]]
                row[field] = str(val).strip() if val is not None else None
            else:
                row[field] = None

        # Fill in from form_meta if not present in row
        if not row.get("customer_name") and form_meta.get("customer_name"):
            row["customer_name"] = form_meta["customer_name"]

        rows.append(row)

    return rows, form_meta


async def sync_sns(rows: list[dict]) -> tuple[int, int]:
    """Sync SNs into sn_assets table. Returns (added_count, skipped_count)."""
    added = 0
    skipped = 0

    async with engine.begin() as conn:
        for row in rows:
            sn = row["sn"]
            customer_code = row.get("customer_code") or "TEST"
            customer_name = row.get("customer_name") or "TestCustomer"
            material_code = row.get("material_code") or "TEST-MAT"
            material_name = row.get("material_name") or "TestMaterial"

            # failure_description is parsed above but sn_assets table has no such column;
            # it is available in row["failure_description"] if needed for reporting.

            # Check if already exists and status is 'valid'
            result = await conn.execute(
                text("SELECT asset_status FROM sn_assets WHERE sn = :sn LIMIT 1"),
                {"sn": sn},
            )
            existing = result.fetchone()
            if existing and existing[0] == "valid":
                print(f"SKIP: SN={sn} already exists (status={existing[0]})")
                skipped += 1
                continue

            # INSERT ... ON DUPLICATE KEY UPDATE for idempotency
            values = {
                "sn": sn,
                "customer_code": customer_code,
                "customer_name": customer_name,
                "material_code": material_code,
                "material_name": material_name,
                "asset_status": "valid",
            }

            columns = ", ".join(values.keys())
            placeholders = ", ".join(f":{k}" for k in values.keys())
            updates = ", ".join(
                f"{k} = VALUES({k})"
                for k in values.keys()
                if k != "sn"
            )

            sql = (
                f"INSERT INTO sn_assets ({columns}) "
                f"VALUES ({placeholders}) "
                f"ON DUPLICATE KEY UPDATE {updates}"
            )

            await conn.execute(text(sql), values)
            print(f"ADDED: SN={sn} customer={customer_name} material={material_code}")
            added += 1

    return added, skipped


async def main() -> None:
    xlsx_path = XLSX_PATH
    if not xlsx_path.exists():
        print(f"ERROR: xlsx file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading xlsx: {xlsx_path}")
    rows, form_meta = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} SN(s) from xlsx")
    if form_meta:
        print(f"Form metadata: {form_meta}")

    if not rows:
        print("No SN rows found, nothing to sync.")
        return

    print("\n--- Syncing to sn_assets ---")
    added, skipped = await sync_sns(rows)
    print(f"\nSummary: Added {added} SNs, Skipped {skipped} SNs")


if __name__ == "__main__":
    asyncio.run(main())
